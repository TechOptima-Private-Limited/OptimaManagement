# assets/export_utils.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.utils import timezone
from .models import (
    AssetType, Asset, AssetAssignment, AssetReturn, 
    AssetHistory, OffboardingAssetReturn, EmployeeStatus, AssetRepair
)


def auto_adjust_column_width(worksheet):
    """Auto-adjust column widths based on content"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width


def style_header_row(worksheet):
    """Apply styling to the header row"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def export_asset_types(worksheet, queryset):
    """Export Asset Types to worksheet"""
    headers = ['ID', 'Name', 'Tag Prefix', 'Category', 'Description', 'Asset Team Email', 'Is Active']
    worksheet.append(headers)
    
    for asset_type in queryset:
        worksheet.append([
            asset_type.id,
            asset_type.name,
            asset_type.tag_prefix,
            asset_type.get_category_display(),
            asset_type.description,
            asset_type.asset_team_email,
            'Yes' if asset_type.is_active else 'No'
        ])
    
    style_header_row(worksheet)
    auto_adjust_column_width(worksheet)


def export_assets(worksheet, queryset):
    """Export Assets to worksheet"""
    headers = [
        'ID', 'Asset Type', 'Name', 'Asset Tag', 'Serial Number', 
        'Status', 'Purchased Date', 'Previously Used By', 'Laptop Age (Days)',
        'Is Active'
    ]
    worksheet.append(headers)
    
    for asset in queryset:
        # Calculate laptop age in days
        laptop_age_days = ''
        if asset.laptop_age:
            laptop_age_days = asset.laptop_age.days
        
        # Get previously used by username
        prev_user = ''
        if asset.previously_used_by:
            prev_user = f"{asset.previously_used_by.get_full_name()} ({asset.previously_used_by.username})"
        
        worksheet.append([
            asset.id,
            asset.asset_type.name,
            asset.name,
            asset.asset_tag,
            asset.serial_number,
            asset.get_status_display(),
            asset.purchased_date.strftime('%Y-%m-%d') if asset.purchased_date else '',
            prev_user,
            laptop_age_days,
            'Yes' if asset.is_active else 'No'
        ])
    
    style_header_row(worksheet)
    auto_adjust_column_width(worksheet)


def export_asset_assignments(worksheet, queryset):
    """Export Asset Assignments to worksheet"""
    headers = [
        'ID', 'Employee', 'Employee Email', 'Assets', 'Asset Types',
        'Manager Email', 'Notes', 'Assigned At', 'Updated At'
    ]
    worksheet.append(headers)
    
    for assignment in queryset:
        # Get employee info
        employee_name = f"{assignment.employee.get_full_name()} ({assignment.employee.username})"
        employee_email = assignment.employee.email
        
        # Get assets list
        assets_list = ', '.join([
            f"{asset.asset_tag} ({asset.name})" 
            for asset in assignment.assets.all()
        ])
        
        # Get asset types list
        asset_types_list = ', '.join([
            asset_type.name 
            for asset_type in assignment.asset_types.all()
        ])
        
        worksheet.append([
            assignment.id,
            employee_name,
            employee_email,
            assets_list,
            asset_types_list,
            assignment.manager_email or '',
            assignment.notes,
            assignment.assigned_at.strftime('%Y-%m-%d %H:%M:%S'),
            assignment.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    style_header_row(worksheet)
    auto_adjust_column_width(worksheet)


def export_asset_returns(worksheet, queryset):
    """Export Asset Returns to worksheet"""
    headers = [
        'ID', 'Assignment ID', 'Employee', 'Asset Tag', 'Asset Name',
        'Condition', 'Notes', 'Returned At'
    ]
    worksheet.append(headers)
    
    for asset_return in queryset:
        employee_name = f"{asset_return.assignment.employee.get_full_name()} ({asset_return.assignment.employee.username})"
        
        worksheet.append([
            asset_return.id,
            asset_return.assignment.id,
            employee_name,
            asset_return.asset.asset_tag,
            asset_return.asset.name,
            asset_return.get_condition_display(),
            asset_return.notes,
            asset_return.returned_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    style_header_row(worksheet)
    auto_adjust_column_width(worksheet)


def export_asset_history(worksheet, queryset):
    """Export Asset History to worksheet"""
    headers = [
        'ID', 'Asset Tag', 'Asset Name', 'Action', 
        'Performed By', 'Performed At', 'Notes'
    ]
    worksheet.append(headers)
    
    for history in queryset:
        performed_by = ''
        if history.performed_by:
            performed_by = f"{history.performed_by.get_full_name()} ({history.performed_by.username})"
        
        worksheet.append([
            history.id,
            history.asset.asset_tag,
            history.asset.name,
            history.action,
            performed_by,
            history.performed_at.strftime('%Y-%m-%d %H:%M:%S'),
            history.notes
        ])
    
    style_header_row(worksheet)
    auto_adjust_column_width(worksheet)


def export_offboarding_returns(worksheet, queryset):
    """Export Offboarding Asset Returns to worksheet"""
    headers = [
        'ID', 'Employee', 'Employee Email', 'Laptop Status', 
        'Returned Assets', 'Remarks', 'Is Offboarded', 
        'Created At', 'Updated At'
    ]
    worksheet.append(headers)
    
    for offboarding in queryset:
        employee_name = f"{offboarding.user.get_full_name()} ({offboarding.user.username})"
        employee_email = offboarding.user.email
        
        # Get returned assets list
        returned_assets = ', '.join([
            f"{asset.asset_tag} ({asset.name})" 
            for asset in offboarding.returned_assets.all()
        ])
        
        worksheet.append([
            offboarding.id,
            employee_name,
            employee_email,
            offboarding.get_laptop_status_display(),
            returned_assets,
            offboarding.remarks or '',
            'Yes' if offboarding.is_offboarded else 'No',
            offboarding.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            offboarding.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    style_header_row(worksheet)
    auto_adjust_column_width(worksheet)


def export_asset_repairs(worksheet, queryset):
    """Export Asset Repairs to worksheet"""
    headers = [
        'ID', 'Asset Tag', 'Asset Name', 'Issue Description', 'Status',
        'Reported By', 'Reported At', 'Started At', 'Completed At',
        'Repair Vendor', 'Repair Cost', 'Estimated Completion', 'Repair Notes'
    ]
    worksheet.append(headers)
    
    for repair in queryset:
        reported_by = ''
        if repair.reported_by:
            reported_by = f"{repair.reported_by.get_full_name()} ({repair.reported_by.username})"
        
        worksheet.append([
            repair.id,
            repair.asset.asset_tag,
            repair.asset.name,
            repair.issue_description,
            repair.get_status_display(),
            reported_by,
            repair.reported_at.strftime('%Y-%m-%d %H:%M:%S'),
            repair.started_at.strftime('%Y-%m-%d %H:%M:%S') if repair.started_at else '',
            repair.completed_at.strftime('%Y-%m-%d %H:%M:%S') if repair.completed_at else '',
            repair.repair_vendor or '',
            str(repair.repair_cost) if repair.repair_cost else '',
            repair.estimated_completion.strftime('%Y-%m-%d') if repair.estimated_completion else '',
            repair.repair_notes
        ])
    
    style_header_row(worksheet)
    auto_adjust_column_width(worksheet)


def export_employee_status(worksheet, queryset):

    """Export Employee Status to worksheet"""
    headers = ['ID', 'Employee', 'Employee Email', 'Is Active']
    worksheet.append(headers)
    
    for emp_status in queryset:
        employee_name = f"{emp_status.employee.get_full_name()} ({emp_status.employee.username})"
        employee_email = emp_status.employee.email
        
        worksheet.append([
            emp_status.id,
            employee_name,
            employee_email,
            'Yes' if emp_status.is_active else 'No'
        ])
    
    style_header_row(worksheet)
    auto_adjust_column_width(worksheet)


def generate_asset_export_excel(user):
    """
    Generate Excel file with all asset management data
    Returns: Workbook object
    """
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Determine user permissions
    admin_groups = ['Admin', 'HR Admin', 'HR Manager', 'IT Support', 'Asset Team']
    is_admin = (
        user.is_superuser 
        or user.is_staff 
        or user.groups.filter(name__in=admin_groups).exists()
    )
    
    # 1. Asset Types
    ws_asset_types = wb.create_sheet("Asset Types")
    if is_admin or user.has_perm('assets.view_assettype'):
        asset_types_qs = AssetType.objects.all().order_by('name')
        if not is_admin:
            asset_types_qs = asset_types_qs.filter(asset_team_email=user.email)
        export_asset_types(ws_asset_types, asset_types_qs)
    else:
        ws_asset_types.append(['No permission to view asset types'])
    
    # 2. Assets
    ws_assets = wb.create_sheet("Assets")
    if is_admin or user.has_perm('assets.view_asset'):
        assets_qs = Asset.objects.select_related('asset_type', 'previously_used_by').order_by('asset_tag')
        if not is_admin:
            asset_type_ids = AssetType.objects.filter(asset_team_email=user.email).values_list('id', flat=True)
            assets_qs = assets_qs.filter(asset_type__id__in=asset_type_ids)
        export_assets(ws_assets, assets_qs)
    else:
        ws_assets.append(['No permission to view assets'])
    
    # 3. Asset Assignments
    ws_assignments = wb.create_sheet("Asset Assignments")
    if is_admin or user.has_perm('assets.view_assetassignment'):
        assignments_qs = AssetAssignment.objects.select_related('employee').prefetch_related('assets', 'asset_types').order_by('-assigned_at')
        if not is_admin:
            assignments_qs = assignments_qs.filter(employee=user)
        export_asset_assignments(ws_assignments, assignments_qs)
    else:
        ws_assignments.append(['No permission to view asset assignments'])
    
    # 4. Asset Returns
    ws_returns = wb.create_sheet("Asset Returns")
    if is_admin or user.has_perm('assets.view_assetreturn'):
        returns_qs = AssetReturn.objects.select_related('assignment__employee', 'asset').order_by('-returned_at')
        if not is_admin:
            returns_qs = returns_qs.filter(assignment__employee=user)
        export_asset_returns(ws_returns, returns_qs)
    else:
        ws_returns.append(['No permission to view asset returns'])
    
    # 5. Asset History
    ws_history = wb.create_sheet("Asset History")
    if is_admin or user.has_perm('assets.view_assethistory'):
        history_qs = AssetHistory.objects.select_related('asset', 'performed_by').order_by('-performed_at')
        if not is_admin:
            history_qs = history_qs.filter(asset__assignments__employee=user).distinct()
        export_asset_history(ws_history, history_qs)
    else:
        ws_history.append(['No permission to view asset history'])
    
    # 6. Offboarding Returns
    ws_offboarding = wb.create_sheet("Offboarding Returns")
    if is_admin or user.has_perm('assets.view_offboardingassetreturn'):
        offboarding_qs = OffboardingAssetReturn.objects.select_related('user').prefetch_related('returned_assets').order_by('-created_at')
        if not is_admin:
            offboarding_qs = offboarding_qs.filter(user=user)
        export_offboarding_returns(ws_offboarding, offboarding_qs)
    else:
        ws_offboarding.append(['No permission to view offboarding returns'])
    
    # 7. Asset Repairs
    ws_repairs = wb.create_sheet("Asset Repairs")
    if is_admin or user.has_perm('assets.view_assetrepair'):
        repairs_qs = AssetRepair.objects.select_related('asset', 'reported_by').order_by('-reported_at')
        if not is_admin:
            repairs_qs = repairs_qs.filter(asset__assignments__employee=user).distinct()
        export_asset_repairs(ws_repairs, repairs_qs)
    else:
        ws_repairs.append(['No permission to view asset repairs'])
    
    # 8. Employee Status
    ws_emp_status = wb.create_sheet("Employee Status")
    if is_admin or user.has_perm('assets.view_employeestatus'):
        emp_status_qs = EmployeeStatus.objects.select_related('employee').order_by('employee__username')
        if not is_admin:
            emp_status_qs = emp_status_qs.filter(employee=user)
        export_employee_status(ws_emp_status, emp_status_qs)
    else:
        ws_emp_status.append(['No permission to view employee status'])
    
    return wb
